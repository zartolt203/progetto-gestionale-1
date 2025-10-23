"""
Main module: defines routes for homepage, authentication and main features.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app
from flask import jsonify, send_file
from flask_login import login_required
from werkzeug.utils import secure_filename
from urllib.parse import quote
from openpyxl import Workbook
from app import db 
from app.models import Item, ItemPictures
import os 
import io
import shutil
import pandas as pd



main_bp = Blueprint('main', __name__)



@main_bp.route('/')
def index():

    """Endpoint for the homepage: retrieves all items from the database, separates them by location, and renders the main template."""

    items = Item.query.all()

    # Separate items by location
    items_magazzino_1   = [i for i in items if i.locazione == 'magazzino-1']
    items_magazzino_2 = [i for i in items if i.locazione == 'magazzino-2']

    # Add picture dictionaries to each item
    for item in items:
        item.picture_dicts = [
            {"file_path": p.file_path, "id": p.id}
            for p in item.pictures
        ]

    return render_template('main.html', items_magazzino_1=items_magazzino_1, items_magazzino_2=items_magazzino_2)



@main_bp.route('/add_item', methods=['POST'])
@login_required
def add_item():

    """Endpoint to add a new item: create an Item object, add it to the database, and redirect to index."""

    # Recover data from the form 
    collo = request.form.get('collo')
    codice = request.form.get('codice')
    descrizione = request.form.get('descrizione')
    quantita = request.form.get('quantita')
    locazione = request.form.get('locazione')
    matricola = request.form.get('matricola')
    note = request.form.get('note')

    # Create a new Item object and add it to the database
    new_item = Item(
        collo=collo,
        codice=codice,
        descrizione=descrizione,
        quantita=quantita,
        locazione=locazione,
        matricola=matricola,
        note=note
    )

    db.session.add(new_item)
    db.session.commit()

    flash('Nuovo collo aggiunto con successo!', 'success')

    return redirect(url_for('main.index'))



@main_bp.route('/delete_item/<int:id>', methods=['POST'])
@login_required
def delete_item(id):

    """Endpoint to delete an item: get item id from user and delete it from database."""

    item = Item.query.get_or_404(id)

    db.session.delete(item)
    db.session.commit()

    # Delete associated folder if it exists
    upload_base = current_app.config['UPLOAD_FOLDER']
    collo_folder = os.path.join(upload_base, secure_filename(item.collo))

    if os.path.exists(collo_folder) and os.path.isdir(collo_folder):
        try:
            shutil.rmtree(collo_folder)  
        except Exception as e:
            return jsonify({'success': False, 'message': f'Errore eliminazione cartella: {str(e)}'}), 500


    return jsonify({'success': True, 'message': 'Elemento eliminato con successo'})



@main_bp.route('/modify_item/<int:id>', methods=['POST'])
@login_required
def modify_item(id):

    """Endpoint to modify an item: get item id from user, update its attributes and save changes to database."""

    item = Item.query.get_or_404(id)

    # Update item attributes with form data
    item.collo = request.form.get('collo', item.collo)
    item.codice = request.form.get('codice', item.codice)
    item.descrizione = request.form.get('descrizione', item.descrizione)
    item.quantita = request.form.get('quantita', item.quantita)
    item.locazione = request.form.get('locazione', item.locazione)
    item.matricola = request.form.get('matricola', item.matricola)
    item.note = request.form.get('note', item.note)

    db.session.commit()

    flash('Collo modificato con successo!', 'success')
    
    return redirect(url_for('main.index'))



@main_bp.route('/request_transfer/<int:id>', methods=['GET'])
@login_required
def request_transfer(id):

    """Opens e-mail client (Outlook) with a precompiled one."""
    
    # Get the item by id
    item = Item.query.get_or_404(id)

    # Fixed recipients list
    recipients = [
        'mail@prova.com'
    ]

    # Determine the departure and destination based on the item's location
    if item.locazione == 'magazzino-1':
        partenza = 'magazzino-1'
        destinazione = 'magazzino-2'
    else:
        partenza = 'magazzino-2'
        destinazione = 'magazzino-1'

    # Prepare the email subject and body
    subject = f"Trasferimento Collo {item.collo}"
    body = (
        f"\nSi richiede trasferimento da {partenza} a {destinazione} di:\n\n"
        f"Collo: {item.collo}\n"
        f"Codice: {item.codice}\n"
        f"Matricola: {item.matricola}\n"
        f"Descrizione: {item.descrizione}\n"
        f"Quantità: {item.quantita}\n"
        f"Note: {item.note}\n\n"
    )

    # URL-encode of subject and body
    mailto = (
        f"mailto:{';'.join(recipients)}"
        f"?subject={quote(subject)}"
        f"&body={quote(body)}"
    )

    return redirect(mailto)



@main_bp.route('/upload_photos', methods=['POST'])
@login_required
def upload_photos():
    
    """Endpoint for uploading photos associated with an item."""

    item_id = request.form.get('item_id')
    files = request.files.getlist('photos')

    if not item_id or not files:
        return jsonify(success=False, error='Dati insufficienti per l\'upload.'), 400

    item = Item.query.get(item_id)

    if not item:
        return jsonify(success=False, error='Elemento non trovato.'), 404

    # Define the upload folder based on the item's collo
    upload_base = current_app.config['UPLOAD_FOLDER']
    collo_folder = os.path.join(upload_base, secure_filename(item.collo))

    # Create the collo folder if it doesn't exist
    os.makedirs(collo_folder, exist_ok=True)

    for file in files:
        if file.filename == '':
            continue # Skip empty files

        filename = secure_filename(file.filename)
        full_path = os.path.join(collo_folder, filename)
        file.save(full_path)

        # Relative path for database storage
        relative_path = os.path.relpath(full_path, start=upload_base).replace('\\', '/')

        picture = ItemPictures(item_id=item.id, file_path=relative_path)
        db.session.add(picture)

    db.session.commit()
    
    # Prepare the JS response with the list of uploaded pictures
    pictures = [
        {"file_path": p.file_path, "id": p.id}
        for p in item.pictures
    ]

    return jsonify(success=True, pictures=pictures)



@main_bp.route('/delete_photos/<int:photo_id>', methods=['POST'])
@login_required
def delete_photos(photo_id):

    """Endpoint for deleting a photo associated with an item."""

    pic = ItemPictures.query.get_or_404(photo_id)

    # Build the full path to the file
    upload_base = current_app.config['UPLOAD_FOLDER']
    full_path = os.path.join(upload_base, pic.file_path)

    # Delete the file from the filesystem
    if os.path.exists(full_path):
        os.remove(full_path)
        flash('Foto eliminata con successo!', 'success')
    else:
        flash('File non trovato sul server.', 'error')
        return redirect(url_for('main.index'))
    
    # Delete the picture record from the database
    db.session.delete(pic)
    db.session.commit()

    # Memorize the parent directory for next check
    parent_dir = os.path.dirname(full_path)

    # If the folder is empty, delete it
    if os.path.isdir(parent_dir) and not os.listdir(parent_dir):
        os.rmdir(parent_dir)

    return ('', 204) 
    


@main_bp.route('/export_xlsx', methods=['POST'])
@login_required
def export_xlsx():
    
    """Endpoint to export a report to an Excel file."""

    # Load data
    items_magazzino_1 = Item.query.filter_by(locazione='magazzino-1').all()
    items_magazzino_2 = Item.query.filter_by(locazione='magazzino-2').all()

    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Riepilogo Colli'

    row = 1

    # Magazzino 1 items
    ws.cell(row, 1, f'Oggetti in magazzino 1: {len(items_magazzino_1)}')
    row += 2
    ws.cell(row, 1, 'Numero collo:')
    ws.cell(row, 2, 'Descrizione:')
    row += 1
    for item in items_magazzino_1:
        ws.cell(row, 1, item.collo)
        ws.cell(row, 2, item.descrizione)
        row += 1

    # Empty rows
    row += 2

    # Magazzino 2 items
    ws.cell(row, 1, f'Oggetti in magazzino 2: {len(items_magazzino_2)}')
    row += 2
    ws.cell(row, 1, 'Numero collo:')
    ws.cell(row, 2, 'Descrizione:')
    row += 1
    for item in items_magazzino_2:
        ws.cell(row, 1, item.collo)
        ws.cell(row, 2, item.descrizione)
        row += 1

    # Expand columns dimensions
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

    # Save in memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the file
    return send_file(
        output,
        as_attachment=True,
        download_name='resocotno_colli_{pd.Timestamp.today().strftime("%d-%m-%Y")}.xlsx',
    )